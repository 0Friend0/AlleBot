from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
import concurrent.futures

class AuctionData():

    data_list = []
    links = []
    counter = 1

    def get_auction_info(self, auction_url):
        # Goes to a provided URL and gets the auction data

        source = requests.get(auction_url).text
        allegro_lokalnie = "allegrolokalnie"
        soup = BeautifulSoup(source, 'lxml')

        if allegro_lokalnie in auction_url:
            auction_name = soup.find('div', class_='offer-action-box').h1.text
            auction_price = soup.find('span', class_='offer-action-box-details__price-value price price--buy-now').text
            seller_name = soup.find('h3', class_='m-b-0').a.text
            seller_comments = "NA"
            return auction_name, auction_price, seller_name, seller_comments
        else:
            auction_name = soup.find('div', class_='_1h7wt _15mod').h1.text
            auction_price = soup.find('div', class_='_1svub _lf05o _9a071_2MEB_').text
            seller_info = soup.find('div', class_='_1h7wt _15mod').a.text
            seller_name = seller_info.split(' ')[0]
            seller_comments = seller_info.split(' ')[-1]
            return auction_name, auction_price, seller_name, seller_comments

    def generate_items_list(self, url):
        # Generate a list of data for every auction

        auction_info = ad.get_auction_info(url)

        item_data = [auction_info[0], auction_info[1], auction_info[2], auction_info[3], url]
        ad.data_list.append(item_data)

    def generate_auction_link_list(self):
        # Creates a list of URL from excel file

        aew = load_workbook("allegro.xlsx")
        aew.active = -1

        while True:
            if aew.active['C' + str(ad.counter + 1)].value is None:
                break
            else:
                link = aew.active['C' + str(ad.counter + 1)].value
                ad.links.append(link)
                ad.counter += 1


        aew.save("allegro.xlsx")


    def add_to_excel(self):
        # Add final auction data to excel


        aew = load_workbook("allegro.xlsx")
        aew.active = -1
        ad.counter = 1

        for item in ad.data_list:

            # Add auction name
            aew.active["A" + str(ad.counter + 1)] = item[0]

            # Add auction price
            aew.active["B" + str(ad.counter + 1)] = item[1]

            # Replace auction URL
            aew.active["C" + str(ad.counter + 1)] = item[4]

            # Add seller name
            aew.active["D" + str(ad.counter + 1)] = item[2]

            # Add seller % comments
            aew.active["E" + str(ad.counter + 1)] = item[3]



            # Increase counter
            ad.counter += 1

            aew.save("allegro.xlsx")

        aew.save("allegro.xlsx")

    def gets_auction_info_master(self):

        ad.generate_auction_link_list()

        with concurrent.futures.ThreadPoolExecutor() as executor:
         executor.map(ad.generate_items_list, ad.links)

        ad.add_to_excel()


ad = AuctionData()

