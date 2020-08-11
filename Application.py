from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook, load_workbook
from chromedriver_path import chromedriverpath
import time
import random
import datetime
from loggingFile import loggingFunction


class Application:
    item_counter = 0
    page_index = 1
    item = 0
    chromedriver_path = chromedriverpath()
    driver = webdriver.Chrome(executable_path=chromedriver_path)
    driver.minimize_window()

    def open(self, item_to_search):

        self.driver = Ap.driver
        self.driver.minimize_window()

        self.driver.maximize_window()
        self.driver.implicitly_wait(5)

        self.driver.get("https://allegro.pl/")

        dialog_box = Ap.check_element_exist("//div [@aria-labelledby = 'dialog-title']")

        # Check if dialog box opens and close it.
        if dialog_box == True:
            self.driver.find_element(By.XPATH, "//div [@aria-labelledby = 'dialog-title']/button").click()

        # Wait random time.
        x = random.uniform(0.1, 0.5)
        time.sleep(x)
        element = self.driver.find_element(By.XPATH,
                                           "//input [@type='search' and contains (@placeholder, 'czego szukasz?')]")

        # Provide item name to Allegro search box with random time between letters.
        for letter in item_to_search:
            x = random.uniform(0.1, 0.5)
            time.sleep(x)
            element.send_keys(letter)

        # Submit and add item name to log file.
        element.submit()
        log.info('Looking for ' + item_to_search)


    def additional_filter_price(self, price_min, price_max):

        # Check if price_min is provided.
        if price_min is not None:
            element = self.driver.find_element(By.XPATH, '//div/input [@id="price_from"]')
            element.send_keys("price_min " + price_min)
            time.sleep(1)

        # Check if price_max is provided.
        if price_max is not None:
            element = self.driver.find_element(By.XPATH, '//div/input [@id="price_to"]')
            element.send_keys("price_max " + price_max)
            time.sleep(1)

        # Check and wait for Allegro finish filtering list of items.
        try:
            element = self.driver.find_element(By.XPATH, '//div [@class="_9c44d_378hD _9c44d_3TOu4"]')
            WebDriverWait(self.driver, 3).until(EC.visibility_of(element))

        except NoSuchElementException:
            pass

        finally:
            element = self.driver.find_element(By.XPATH, '//div [@class="_9c44d_378hD"]')
            element = WebDriverWait(self.driver, 3).until(EC.visibility_of(element))




    def create_excel_file(self, item_to_search):

        # Create excel file and add new tab.
        today = datetime.datetime.now()
        today_fix = today.strftime("%d %m %Y")

        try:
            aew = load_workbook(item_to_search + ".xlsx")
            ws = aew.create_sheet(today_fix)

        except FileNotFoundError:
            aew = Workbook()
            ws = aew.create_sheet(today_fix)

        ws['A1'] = "Auction name"
        ws['B1'] = "Price"
        ws['C1'] = "Auction link"
        ws['D1'] = "Seller name"
        ws['E1'] = "Positive comments"
        aew.save(item_to_search + ".xlsx")

    def finding_items(self, item_to_search):
        aew = load_workbook(item_to_search + ".xlsx")
        last_sheet = aew.sheetnames[-1]
        ws = aew[last_sheet]
        Ap.item = 0

        # Get links for each item on the page.
        for _ in self.driver.find_elements(By.XPATH, "//article [@data-analytics-view-custom-page = '" +
                                                        str(Ap.page_index) + "']"):
            log.debug('Checking the ' + str(Ap.item_counter + 1) + ' item.')

            try:

                # Getting link to item
                item_link = Ap.get_link()
                h = ws["C" + str(Ap.item_counter + 2)]
                h.value = item_link
            except NoSuchElementException:
                continue
            Ap.item_counter += 1
            Ap.item += 1
        aew.save(item_to_search + ".xlsx")

        # Checking if next page of items is available.
        if Ap.check_next_page():
            Ap.finding_items(item_to_search)

    def driver_close(self):
        # Close WebDriver.
        log.info('Found ' + str(Ap.item_counter) + ' items.')
        print(Ap.item_counter)
        self.driver.close()

    def check_element_exist(self, xpath):
        try:
            self.driver.find_element(By.XPATH, xpath)
        except NoSuchElementException:
            return False
        else:
            return True

    def check_next_page(self):
        # Check if next page of items is available.
        page = self.driver.find_element(By.XPATH, "//div [@data-prototype-id='allegro.pagination']/div/div/div/span")
        actions = ActionChains(self.driver)
        actions.move_to_element(page).perform()

        page_count = page.get_attribute("textContent")
        if Ap.page_index < int(page_count):
            element = self.driver.find_element(By.XPATH,
                                               "//div [@data-prototype-id='allegro.pagination']/div/div/a [@data-role='next-page']")
            element.click()
            Ap.page_index += 1
            return True
        else:
            return False

    def get_link(self):
        # Get auction link to item.
        element = self.driver.find_element(By.XPATH,
                                           "//article [@data-analytics-view-custom-page = '" + str(
                                               Ap.page_index) + "' and @data-analytics-view-custom-index0 = '" + str(
                                               Ap.item) + "']/div/div/div[1]/h2/a")
        href = element.get_attribute("href")
        return str(href)


log = loggingFunction()
Ap = Application()

